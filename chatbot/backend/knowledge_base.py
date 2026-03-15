"""
Knowledge base: create by vendor name, ingest PDF/Excel with chunking,
OpenAI embeddings; MongoDB for KB entity and chunks (including embeddings).
Hybrid search: MongoDB $vectorSearch + in-memory BM25, RRF merge.
"""
import io
import logging
import re
import uuid
from typing import Any, Callable

from langchain_openai import OpenAIEmbeddings
from PyPDF2 import PdfReader
from rank_bm25 import BM25Okapi

from config import OPENAI_API_KEY, OPENAI_BASE_URL

logger = logging.getLogger("chatbot.kb")

# OpenAI text-embedding-3-small dimension
EMBEDDING_DIM = 1536


class DuplicateVendorError(Exception):
    def __init__(self, message: str, existing_id: str):
        super().__init__(message)
        self.message = message
        self.existing_id = existing_id


class KBEntry:
    """Lightweight KB entity (from MongoDB). Used as return type for get_kb."""

    __slots__ = ("kb_id", "vendor_name")

    def __init__(self, kb_id: str, vendor_name: str):
        self.kb_id = kb_id
        self.vendor_name = vendor_name


def _normalize_vendor_name(name: str) -> str:
    return (name or "").strip().lower()


def _tokenize_for_bm25(text: str) -> list[str]:
    """Simple tokenizer: lowercase, alphanumeric tokens."""
    text = (text or "").lower()
    tokens = re.findall(r"[a-z0-9]+", text)
    return tokens


# --- Chunking ---


def _chunk_pdf(file_bytes: bytes, source_filename: str) -> list[tuple[str, dict[str, Any]]]:
    """Chunk PDF by page. Returns list of (text, metadata)."""
    reader = PdfReader(io.BytesIO(file_bytes))
    chunks: list[tuple[str, dict[str, Any]]] = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            chunks.append((text.strip(), {"source_filename": source_filename, "page": i + 1}))
    logger.info("PDF %s: %d pages -> %d chunks", source_filename, len(reader.pages), len(chunks))
    return chunks


def _chunk_excel(file_bytes: bytes, source_filename: str) -> list[tuple[str, dict[str, Any]]]:
    """Chunk Excel by row (one chunk per row). Uses openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    chunks: list[tuple[str, dict[str, Any]]] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            parts = [str(c) if c is not None else "" for c in row]
            text = " | ".join(parts).strip()
            if text:
                chunks.append((text, {"source_filename": source_filename, "sheet": sheet_name, "row": row_idx}))
    wb.close()
    logger.info("Excel %s: %d chunks", source_filename, len(chunks))
    return chunks


def chunk_file(file_bytes: bytes, source_filename: str) -> list[tuple[str, dict[str, Any]]]:
    """Dispatch by extension. Returns list of (text, metadata). Raises ValueError for unsupported type."""
    ext = (source_filename or "").rsplit(".", 1)[-1].lower() if "." in (source_filename or "") else ""
    if ext == "pdf":
        return _chunk_pdf(file_bytes, source_filename)
    if ext in ("xlsx", "xls"):
        return _chunk_excel(file_bytes, source_filename)
    raise ValueError(f"Unsupported file type for knowledge base: {ext or 'unknown'}. Use .pdf or .xlsx.")


# --- Embeddings ---

_embeddings_model: OpenAIEmbeddings | None = None


def _get_embeddings() -> OpenAIEmbeddings:
    global _embeddings_model
    if _embeddings_model is None:
        kwargs: dict[str, Any] = {"model": "text-embedding-3-small"}
        if OPENAI_API_KEY:
            kwargs["api_key"] = OPENAI_API_KEY
        if OPENAI_BASE_URL:
            kwargs["base_url"] = OPENAI_BASE_URL
        _embeddings_model = OpenAIEmbeddings(**kwargs)
    return _embeddings_model


def embed_texts(texts: list[str]) -> list[list[float]]:
    """Embed a list of texts with OpenAI. Returns list of embedding vectors."""
    if not texts:
        return []
    emb = _get_embeddings()
    return emb.embed_documents(texts)


# --- KB CRUD (async, MongoDB) ---


async def create_knowledge_base(vendor_name: str) -> tuple[str, str]:
    """
    Create a new KB. Returns (kb_id, vendor_name).
    Raises DuplicateVendorError if vendor already exists.
    """
    import db as db_service
    normalized = _normalize_vendor_name(vendor_name)
    if not normalized:
        raise ValueError("Vendor name is required")
    existing = await db_service.knowledge_bases_col.find_one({"normalized_vendor_name": normalized})
    if existing:
        raise DuplicateVendorError(f"Vendor already exists: {vendor_name}", existing["_id"])
    kb_id = str(uuid.uuid4())
    doc = {
        "_id": kb_id,
        "vendor_name": vendor_name.strip(),
        "normalized_vendor_name": normalized,
    }
    await db_service.knowledge_bases_col.insert_one(doc)
    logger.info("Created KB %s for vendor %s", kb_id[:8], vendor_name)
    return kb_id, doc["vendor_name"]


async def check_duplicate(vendor_name: str) -> tuple[bool, str | None]:
    """Returns (exists, existing_kb_id or None)."""
    import db as db_service
    normalized = _normalize_vendor_name(vendor_name)
    if not normalized:
        return False, None
    doc = await db_service.knowledge_bases_col.find_one({"normalized_vendor_name": normalized})
    return (True, doc["_id"]) if doc else (False, None)


async def list_knowledge_bases() -> list[dict[str, str]]:
    """Returns [{ id, vendor_name }]."""
    import db as db_service
    cursor = db_service.knowledge_bases_col.find({}, {"_id": 1, "vendor_name": 1})
    return [{"id": doc["_id"], "vendor_name": doc["vendor_name"]} async for doc in cursor]


async def get_kb_id_by_vendor_name(vendor_name: str) -> str | None:
    """
    Resolve vendor name to a KB id. Prefer exact match on normalized_vendor_name;
    else first KB whose normalized_vendor_name contains the normalized input (e.g. "Glosil" -> "Glosil Scientific").
    """
    import db as db_service
    normalized = _normalize_vendor_name(vendor_name)
    if not normalized:
        return None
    doc = await db_service.knowledge_bases_col.find_one({"normalized_vendor_name": normalized})
    if doc:
        return doc["_id"]
    cursor = db_service.knowledge_bases_col.find({}, {"_id": 1, "normalized_vendor_name": 1})
    async for d in cursor:
        if normalized in (d.get("normalized_vendor_name") or ""):
            return d["_id"]
    return None


async def get_kb(kb_id: str) -> KBEntry | None:
    """Return KB entity or None."""
    import db as db_service
    doc = await db_service.knowledge_bases_col.find_one({"_id": kb_id})
    if not doc:
        return None
    return KBEntry(kb_id=doc["_id"], vendor_name=doc["vendor_name"])


async def get_documents(kb_id: str) -> list[dict[str, Any]]:
    """Aggregate by source_filename -> chunk_count. Returns [{ source_filename, chunk_count }]."""
    import db as db_service
    cursor = db_service.kb_chunks_col.aggregate([
        {"$match": {"kb_id": kb_id}},
        {"$group": {"_id": "$source_filename", "chunk_count": {"$sum": 1}}},
        {"$project": {"_id": 0, "source_filename": "$_id", "chunk_count": 1}},
    ])
    return [doc async for doc in cursor]


async def delete_knowledge_base(kb_id: str) -> bool:
    """Remove KB and all its chunks from MongoDB. Returns True if removed."""
    import db as db_service
    doc = await db_service.knowledge_bases_col.find_one_and_delete({"_id": kb_id})
    if not doc:
        return False
    await db_service.kb_chunks_col.delete_many({"kb_id": kb_id})
    logger.info("Deleted KB %s", kb_id[:8])
    return True


async def remove_documents(kb_id: str, source_filenames: list[str]) -> tuple[int, int]:
    """
    Remove all chunks whose source_filename is in the list.
    Returns (files_removed, chunks_deleted).
    """
    import db as db_service
    entry = await get_kb(kb_id)
    if not entry:
        raise KeyError(f"Knowledge base not found: {kb_id}")
    to_remove = set(f.strip() for f in source_filenames if f and f.strip())
    if not to_remove:
        return 0, 0
    result = await db_service.kb_chunks_col.delete_many({
        "kb_id": kb_id,
        "source_filename": {"$in": list(to_remove)},
    })
    deleted = result.deleted_count
    logger.info("KB %s: removed %d files, %d chunks", kb_id[:8], len(to_remove), deleted)
    return len(to_remove), deleted


async def add_chunks_to_kb(
    kb_id: str,
    chunks: list[tuple[str, dict[str, Any]]],
    embed_batch_size: int = 50,
    progress_callback: Callable[[int, int], None] | None = None,
) -> int:
    """
    Append chunks (text, metadata) to KB: embed and insert into MongoDB.
    progress_callback(current, total) optional. Returns number of chunks added.
    """
    entry = await get_kb(kb_id)
    if not entry:
        raise KeyError(f"Knowledge base not found: {kb_id}")
    if not chunks:
        return 0
    all_texts = [c[0] for c in chunks]
    all_embeddings: list[list[float]] = []
    total = len(all_texts)
    for i in range(0, total, embed_batch_size):
        batch = all_texts[i : i + embed_batch_size]
        all_embeddings.extend(embed_texts(batch))
        if progress_callback:
            progress_callback(min(i + len(batch), total), total)
    triples: list[tuple[str, dict[str, Any], list[float]]] = []
    for j, (text, meta) in enumerate(chunks):
        emb = all_embeddings[j] if j < len(all_embeddings) else []
        triples.append((text, meta, emb))
    return await add_chunks_with_embeddings(kb_id, triples)


async def add_chunks_with_embeddings(
    kb_id: str,
    chunks: list[tuple[str, dict[str, Any], list[float]]],
) -> int:
    """Insert pre-embedded chunks into MongoDB (kb_chunks). Returns count added."""
    import db as db_service
    entry = await get_kb(kb_id)
    if not entry:
        raise KeyError(f"Knowledge base not found: {kb_id}")
    if not chunks:
        return 0
    docs = []
    for text, meta, emb in chunks:
        chunk_id = str(uuid.uuid4())
        docs.append({
            "_id": chunk_id,
            "kb_id": kb_id,
            "text": text,
            "source_filename": meta.get("source_filename") or "unknown",
            "metadata": meta,
            "embedding": emb,
        })
    await db_service.kb_chunks_col.insert_many(docs)
    logger.info("Inserted %d chunks into KB %s", len(docs), kb_id[:8])
    return len(docs)


# --- Hybrid search (MongoDB $vectorSearch + BM25, RRF) ---

RRF_K = 60
DEFAULT_TOP_K = 10

# Name of the vector search index (must match db.ensure_kb_indexes)
KB_CHUNKS_VECTOR_INDEX = "kb_chunks_vector"


async def hybrid_search(kb_id: str, query: str, top_k: int = DEFAULT_TOP_K) -> list[str]:
    """
    Run hybrid search: vector top-k via MongoDB $vectorSearch, BM25 top-k from chunk texts in MongoDB, RRF merge.
    Returns list of chunk texts (up to top_k).
    """
    import db as db_service
    entry = await get_kb(kb_id)
    if not entry:
        return []

    query_emb = embed_texts([query])[0]
    vec_top_texts: list[str] = []

    # Vector search via $vectorSearch
    try:
        pipeline = [
            {
                "$vectorSearch": {
                    "index": KB_CHUNKS_VECTOR_INDEX,
                    "path": "embedding",
                    "queryVector": query_emb,
                    "numSearch": top_k * 2,
                    "filter": {"kb_id": kb_id},
                }
            },
            {"$project": {"text": 1, "_id": 0}},
            {"$limit": top_k * 2},
        ]
        cursor = db_service.kb_chunks_col.aggregate(pipeline)
        async for doc in cursor:
            if doc.get("text"):
                vec_top_texts.append(doc["text"])
    except Exception as e:
        logger.warning("MongoDB $vectorSearch failed for kb_id=%s: %s", kb_id[:8], e)

    # BM25: load all chunk texts for this KB from MongoDB
    bm25_top_texts: list[str] = []
    cursor = db_service.kb_chunks_col.find({"kb_id": kb_id}, {"text": 1})
    chunk_texts: list[str] = []
    async for doc in cursor:
        if doc.get("text"):
            chunk_texts.append(doc["text"])
    if chunk_texts and query:
        tokenized = [_tokenize_for_bm25(t) for t in chunk_texts]
        bm25 = BM25Okapi(tokenized)
        tokenized_query = _tokenize_for_bm25(query)
        if tokenized_query:
            scores = bm25.get_scores(tokenized_query)
            ranked = sorted(range(len(scores)), key=lambda i: -scores[i])
            for i in ranked[: top_k * 2]:
                if scores[i] > 0 and chunk_texts[i]:
                    bm25_top_texts.append(chunk_texts[i])

    # RRF merge by text
    rrf_scores: dict[str, float] = {}
    for rank, text in enumerate(vec_top_texts, start=1):
        if text:
            rrf_scores[text] = rrf_scores.get(text, 0) + 1.0 / (RRF_K + rank)
    for rank, text in enumerate(bm25_top_texts, start=1):
        if text:
            rrf_scores[text] = rrf_scores.get(text, 0) + 1.0 / (RRF_K + rank)

    merged = sorted(rrf_scores.items(), key=lambda x: -x[1])[:top_k]
    return [text for text, _ in merged]


async def get_kb_mongo_status() -> dict[str, Any]:
    """Return MongoDB KB status: collection names and document counts. Replaces get_qdrant_status."""
    import db as db_service
    try:
        kb_count = await db_service.knowledge_bases_col.count_documents({})
        chunk_count = await db_service.kb_chunks_col.count_documents({})
        return {
            "enabled": True,
            "knowledge_bases_count": kb_count,
            "kb_chunks_count": chunk_count,
        }
    except Exception as e:
        return {"enabled": True, "error": str(e)}
